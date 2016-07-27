import openpyxl
import webbrowser
import shutil, os, re
import time
import smtplib
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC # available since 2.26.0


path = '/home/pi/working/'
os.chdir(path)
filelist = os.listdir(path)[0]

wb = openpyxl.load_workbook(filelist)
names =wb.get_sheet_names()


if "FinishQuestions" in names:
    print "Processing FinishQuestions"
else:
    print "exiting"
    exit()



sheet = wb.get_sheet_by_name('FinishQuestions')
user = sheet['B1'].value
code= str(sheet['B2'].value)



browser=webdriver.Firefox()
browser.get('https://work.fleetmatics.com/')
#browser.maximize_window()
emailElem = browser.find_element_by_id('USER_NAME')
emailElem.send_keys(user)
print user
passwordElem = browser.find_element_by_id('PASSWORD')
print code
passwordElem.send_keys(code)
button = browser.find_element_by_id('ImageButton1').click()

browser.get('https://work.fleetmatics.com/Admin/Options/JobFinishQuestionList.aspx')

#begin Loop
r=4

while user != 'None':
    button = browser.find_element_by_id('ctl00_BodyContentPlaceHolder_btnAdd').click()
    # add question Name
    Elem = browser.find_element_by_id('ctl00_BodyContentPlaceHolder_txtJobFinishQuestion')
    Elem.send_keys(sheet.cell(row=r, column=1).value)
    # add question ID number
    Elem1 = browser.find_element_by_id('ctl00_BodyContentPlaceHolder_txtFinishQuestionId')
    Elem1 = browser.find_element_by_id('ctl00_BodyContentPlaceHolder_txtFinishQuestionId').clear()
    print sheet.cell(row=r, column=2).value
    user = str(sheet.cell(row=r, column=2).value)
    Elem1 = browser.find_element_by_id('ctl00_BodyContentPlaceHolder_txtFinishQuestionId')
    Elem1.send_keys(user)
    # add question Job Type
    user = str(sheet.cell(row=r, column=3).value)
    select = Select(browser.find_element_by_id("ctl00_BodyContentPlaceHolder_lstJobtype"))
    select.select_by_visible_text(user)
    print user
    # add question Question Type
    user = str(sheet.cell(row=r, column=4).value)
    print user
    select = Select(browser.find_element_by_id("ctl00_BodyContentPlaceHolder_drpQuestionType"))
    select.select_by_visible_text(user)
    # add Answer if necessary
    c=5
    user = str(sheet.cell(row=r, column=c).value)
    while user != 'None':
        user = str(sheet.cell(row=r, column=c).value)
        if user=='None':
            break
        print user, r, c
        #
        time.sleep(1)
        element = WebDriverWait(browser, 50).until(EC.presence_of_element_located((By.ID, "ctl00_BodyContentPlaceHolder_txtAnswerOption")))
        time.sleep(1)
        Elem1 = browser.find_element_by_id('ctl00_BodyContentPlaceHolder_txtAnswerOption').clear()
        print "Clear"
        element = WebDriverWait(browser, 50).until(EC.presence_of_element_located((By.ID, "ctl00_BodyContentPlaceHolder_txtAnswerOption")))
        time.sleep(1)
        Elem1 = browser.find_element_by_id('ctl00_BodyContentPlaceHolder_txtAnswerOption')
        print "about to send keys"
        element = WebDriverWait(browser, 50).until(EC.presence_of_element_located((By.ID, "ctl00_BodyContentPlaceHolder_txtAnswerOption")))
        time.sleep(1)
        print "Sending keys: ",user
        element = WebDriverWait(browser, 50).until(EC.presence_of_element_located((By.ID, "ctl00_BodyContentPlaceHolder_txtAnswerOption")))
        time.sleep(1)
        Elem1.send_keys(user)
        print "Sent keys"
        time.sleep(1)
        element = WebDriverWait(browser, 50).until(EC.presence_of_element_located((By.ID, "ctl00_BodyContentPlaceHolder_btnAdd")))
        time.sleep(1)
        button = browser.find_element_by_id('ctl00_BodyContentPlaceHolder_btnAdd').click()
        c=c+1
        print "loop"
        time.sleep(1)
        user = str(sheet.cell(row=r, column=c).value)
#    browser.implicitly_wait(1)
    element = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.ID, "ctl00_PageNameImagePlaceHolder_img_btnsave")))
    print "Saving"
    button = browser.find_element_by_id('ctl00_PageNameImagePlaceHolder_img_btnsave').click()
    r=r+1
    user = str(sheet.cell(row=r, column=1).value)

browser.get('https://work.fleetmatics.com/Admin/Logout.aspx')
print "done"
browser.close()

filelist = os.listdir(path)[0]
asfile = '/home/pi/done/' + time.ctime() + filelist + '.xlsx'
shutil.move(filelist, asfile)

server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login("david.balla.fleetmatics@gmail.com", "Fleet2015")
 
msg = "FinishQuestions Complete " + asfile
server.sendmail("david.balla.fleetmatics@gmail.com", "david.balla@fleetmatics.com", msg)
server.quit()
